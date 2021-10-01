import os, re, random, math
from django import template
from django.db.models import query
from django.db.models.query_utils import Q
from django.shortcuts import render
from django.http.response import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from django.conf import settings
from numpy import object_

# Rest Frameworks
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status
from django.core.paginator import Paginator

from openpyxl import * 
from openpyxl.writer.excel import save_virtual_workbook
import pandas as pd
from openpyxl import load_workbook

# Models
from api.models import Book, Seller, Recommend, ExcelFileUpload, UserType, Order

# Create your views here.

def home_index(request):
    return redirect('api/')

def index(request):
    return HttpResponse("Backend Connected")

@api_view(['POST'])
@permission_classes([AllowAny,])
@csrf_exempt
def filterByStalls(request):
    valid_query_parameters = [ 'qid', 'page_number', 'filter_by_subject', 'medium', 'sort_by', 'order_by' ]
    valid_sort_paramters = ['price', 'year_of_publication']

    if request.method == 'POST':
        query = dict( request.data )
        query_parameters = list( query.keys() )
        print(query)

        if 'qid' not in query_parameters or 'page_number' not in query_parameters:
            return Response({
                'status' : 'error',
                'code' : '400',
                'message' : 'qid or page number missing'
            }, status = status.HTTP_200_OK)
        
        for key in query_parameters:
            if key not in valid_query_parameters:
                return Response({
                    'status' : 'error', 
                    'code' : '404',
                    'message' : f'Query is not valid: {key}',
                }, status = status.HTTP_200_OK)
        
        if 'sort_by' in query_parameters and query['sort_by'] not in valid_sort_paramters:
            return Response({
                    'status' : 'error', 
                    'code' : '404',
                    'message' : f"Sort By is not valid: {query['sort_by']}",
                }, status = status.HTTP_200_OK)

        if 'sort_by' in query_parameters and 'order_by' not in query_parameters:
            return Response({
                    'status' : 'error', 
                    'code' : '404',
                    'message' : f"Order By is to passed",
                }, status = status.HTTP_200_OK)

        qid = int(query['qid'])
        page_number = int(query['page_number'])

        if qid not in range(1, 10):
            return Response({
                    'status' : 'error', 
                    'code' : '404',
                    'message' : f"No such seller with this ID exists",
                }, status = status.HTTP_200_OK)

        if 'filter_by_subject' in query_parameters and 'medium' in query_parameters:
            books = Book.objects.filter( seller_id = query['qid'], subject__icontains = query['filter_by_subject'], medium__icontains = query['medium'] ).order_by('id')
        elif 'medium' in query_parameters:
            books = Book.objects.filter( seller_id = query['qid'], medium__icontains = query['medium'] ).order_by('id')
        elif 'filter_by_subject' in query_parameters:
            books = Book.objects.filter( seller_id = query['qid'], subject__icontains = query['filter_by_subject'] ).order_by('id')
        else:
            books = Book.objects.filter(seller_id = query['qid'] ).order_by('id')
        
        if 'sort_by' in query_parameters and query['order_by'] == 'desc':
            books = books.order_by( '-' + str(query['sort_by']))
        elif 'sort_by' in query_parameters:
            books = books.order_by(query['sort_by'])

        listBooks = Paginator(books.values(), 8)
        count = listBooks.count

        if count == 0:
            return Response({
                    'status':"success",
                    'code':200,
                    'message':"Nothing found",
                }, status = status.HTTP_200_OK)
        
        total_page = math.ceil( count / 8 )

        if page_number > total_page:
            return Response({
                'status':"error",
                'code':404,
                'message':"Page Number Exceeded permissible range",
            }, status = status.HTTP_200_OK)

        page = listBooks.page(page_number)

        context = {
            'status':'success',
            'code':200,
            'count': count, 
            'total_page':math.ceil( count / 8 ),
            'current_page':query['page_number'],
            'total_data': len(list(page)) ,
            'data':list(page),
        }

        return Response(context, status = status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([AllowAny,])
@csrf_exempt
def filterBooks(request):
    valid_query_parameters = [ 'q', 'page_number', 'filter_by_subject', 'medium', 'sort_by', 'order_by' ]
    valid_sort_paramters = ['price', 'year_of_publication']

    if request.method == 'POST':
        query = dict( request.data )
        query_parameters = list( query.keys() )
        print(query)

        if 'q' not in query_parameters or 'page_number' not in query_parameters:
            return Response({
                'status' : 'error',
                'code' : '400',
                'message' : 'q or page number missing'
            }, status = status.HTTP_200_OK)
        
        for key in query_parameters:
            if key not in valid_query_parameters:
                return Response({
                    'status' : 'error', 
                    'code' : '404',
                    'message' : f'Query is not valid: {key}',
                }, status = status.HTTP_200_OK)
        
        if 'sort_by' in query_parameters and query['sort_by'] not in valid_sort_paramters:
            return Response({
                    'status' : 'error', 
                    'code' : '404',
                    'message' : f"Sort By is not valid: {query['sort_by']}",
                }, status = status.HTTP_200_OK)

        if 'sort_by' in query_parameters and 'order_by' not in query_parameters:
            return Response({
                    'status' : 'error', 
                    'code' : '404',
                    'message' : f"Order By is to passed",
                }, status = status.HTTP_200_OK)

        q = query['q']
        page_number = int(query['page_number'])

        try:
            books = Book.objects.filter(subject__icontains = q) | Book.objects.filter( author__icontains = q ) | Book.objects.filter( title__icontains = q )
        
        except:
                return Response({
                    'status':"error",
                    'code':404,
                    'message': 'Something went wrong',
                }, status = status.HTTP_200_OK)

        if 'filter_by_subject' in query_parameters and 'medium' in query_parameters:
            books = books.filter( subject__icontains = query['filter_by_subject'], medium__icontains = query['medium'] ).order_by('id')
        elif 'medium' in query_parameters:
            books = books.filter( medium__icontains = query['medium'] ).order_by('id')
        elif 'filter_by_subject' in query_parameters:
            books = books.filter( subject__icontains = query['filter_by_subject'] ).order_by('id')
        
        if 'sort_by' in query_parameters and query['order_by'] == 'desc':
            books = books.order_by( '-' + str(query['sort_by']))
        elif 'sort_by' in query_parameters:
            books = books.order_by(query['sort_by'])
        
        listBooks = Paginator(books.values(), 8)
        count = listBooks.count

        if count == 0:
            return Response({
                    'status':"success",
                    'code':200,
                    'message':"Nothing found",
                }, status = status.HTTP_200_OK)
        
        total_page = math.ceil( count / 8 )

        if page_number > total_page:
            return Response({
                'status':"error",
                'code':404,
                'message':"Page Number Exceeded permissible range",
            }, status = status.HTTP_200_OK)

        page = listBooks.page(page_number)

        context = {
            'status':'success',
            'code':200,
            'count': count, 
            'total_page':math.ceil( count / 8 ),
            'current_page':query['page_number'],
            'total_data': len(list(page)) ,
            'data':list(page),
        }

        return Response(context, status = status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes((AllowAny,))
def getBook(request, book_id):
    if book_id is None:
        return JsonResponse({
            "status" : "error",
            "code" : 404,
            "message" : "No Such Book Available"
        })
    try:
        book = Book.objects.filter(id=book_id)
        if len(book.values()) == 0:
            return Response({
                "error":"404 NOT Found",
                'message':f"Book Id does not exit ${book_id}"
            }, status = status.HTTP_404_NOT_FOUND)
    except:
        return Response({
            'error':"404 NOT FOUND",
            'message': "Book Id "+str(book_id) + " does not exist"
        }, status = status.HTTP_404_NOT_FOUND)
    return Response({'data': book.values()})


@api_view(['POST'])
@permission_classes((AllowAny,))
def userType(request):
    if request.method == 'POST':
        admins = ['acquisition.library', 'pahmad', 'shweta.pandey', 'librarian']
        data = request.data
        regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        if re.fullmatch(regex, data['email']):
            email = data['email']
            username, domain = email.split('@')
            if domain == 'lnmiit.ac.in':
                if username in admins:
                    return Response({
                        "status" : 'Success',
                        "code" : 200,
                        "Message" : 'User is admin', 
                        'data' : {
                            'type':'admin',
                            'code' : 1
                        }
                    }, status = status.HTTP_200_OK)
                
                elif username.isalnum() and username[2] in ['p', 'm']:
                    return Response({
                        'status' : 'Success',
                        'code' : 200,
                        'message' : 'User is Research Scholar',
                        'data' : {
                            'type' : 'Scholar', 
                            'code' : 3
                        }
                    }, status=status.HTTP_200_OK)

                elif not username.isalnum():
                    return JsonResponse({
                        "status" : 'Success',
                        "code" : 200,
                        "Message" : 'User is Professor', 
                        'data' : {
                            'type':'Professor',
                            'code' : 2
                        }
                    })
            else:
                return Response({
                    "status" : 'Error',
                    'code': 403,
                    'Message' : " Enter a Valid LNMIIT Email only"
                }, status=status.HTTP_403_FORBIDDEN)
        else:
            return Response({
                "status" : 'Error',
                'code': 403,
                'Message' : " Enter a Valid Email "
            }, status=status.HTTP_403_FORBIDDEN)

        return Response({
            "status" : 'Success',
            "code" : 200,
            "Message" : 'User is Student', 
            'data' : {
                'type':'Student',
                'code' : 4
            }
        }, status=status.HTTP_202_ACCEPTED)


from django.template.loader import render_to_string, get_template
from django.core.mail import EmailMessage

@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def recommendApi(request):
    valid_keys = ['items', 'email', 'name', 'type', 'code']
    query = dict( request.data )
    query_parameters = list( query.keys() )

    for key in valid_keys:
        if key not in query_parameters:
            return Response({
                'status' : "Error",
                'code' : 400,
                'message' : f"Wrong Query Passed {key}"
            }, status=status.HTTP_206_PARTIAL_CONTENT)
    
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email = query['email']
    if re.fullmatch(regex, email):
        username, domain = email.split('@')
        if domain != 'lnmiit.ac.in':
            return Response({
                'status' : "Error",
                'code' : 206,
                'message' : f"Enter a Valid LNMIIT Email Only"
            }, status = status.HTTP_206_PARTIAL_CONTENT)

        items = query['items']
        name = query['name']
        code = query['code']
        books = []
        userType = UserType.objects.get( code = code )
        row = 1
        for item in items:

            book = Book.objects.get( id = item['id'] )
            books.append( { 'id' : row, 'title' : book.title, 'author' : book.author, 'subject' : book.subject, 'publisher' : book.publisher, 'medium' : book.medium, 'demo' : book.price_denomination, 'price' : book. expected_price, 'quantity' : item['qty'] } )
            recommend = Recommend()
            recommend.book = book
            recommend.quantity = item['qty']
            recommend.name = name
            recommend.email = email
            recommend.usertype = userType
            recommend.recommended_to_library = True
            row += 1
            recommend.save()
        ctx = {
            'name' : name,
            'total' : len(books),
            'books' : books
        }
        template = get_template( 'email.html').render(ctx)

        msg = EmailMessage(
            'Thanks For Recommending Books to the Library',
            template,
            settings.EMAIL_HOST_USER,
            [email]
        )
        
        msg.content_subtype = "html" 
        msg.send()
        return Response({
            'status' : "Success",
            'code' : 201,
            'message' : "Added Recommendations",
        }, status = status.HTTP_201_CREATED)

    else:
        return Response({
            'status' : "Error",
            'code' : 206,
            'message' : f"Enter a Valid Email"
        }, status=status.HTTP_206_PARTIAL_CONTENT)


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def purchaseApi(request):
    valid_keys = ['email', 'name', 'type', 'code', 'book_id', 'qty']
    query = dict( request.data )
    query_parameters = list( query.keys() )

    for key in valid_keys:
        if key not in query_parameters:
            return Response({
                'status' : "Error",
                'code' : 400,
                'message' : f"Wrong Query Passed {key}"
            }, status=status.HTTP_206_PARTIAL_CONTENT)
    
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email = query['email']

    if re.fullmatch(regex, email):
        username, domain = email.split('@')
        if domain != 'lnmiit.ac.in':
            return Response({
                'status' : "Error",
                'code' : 206,
                'message' : f"Enter a Valid LNMIIT Email Only"
            }, status = status.HTTP_206_PARTIAL_CONTENT)
        name = query['name']
        code = query['code']
        book_id = query['book_id']
        qty = query['qty']
        userType = UserType.objects.get( code = code )
        if userType.code == 4:
            return Response({
                'status' : 'Error',
                'code' : 400,
                'message' : "Students are allowed to do Personal Purchase"
            }, status= status.HTTP_200_OK)
        try:
            book = Book.objects.get( id = book_id )
            order = Order()
            order.email = email
            order.name = name
            order.book = book
            order.quantity = qty
            order.is_ordered = True
            order.usertype = userType

            order.save()

            print(f'Book Ordered with ID: {order.id}-{order.quantity}')
            
            return Response({
                'status' : "Success",
                'code' : 201,
                'message' : "Book Ordered Success",
            }, status = status.HTTP_201_CREATED)
        except:
            return Response({
                'status' : 'Error', 
                'code' : 400,
                'message' : "Something Wrong with order table"
            }, status = status.HTTP_400_BAD_REQUEST)



    else:
        return Response({
            'status' : "Error",
            'code' : 206,
            'message' : f"Enter a Valid Email"
        }, status=status.HTTP_206_PARTIAL_CONTENT)


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def getRecommendation(request):
    valid_keys = ['email', 'name']
    query = dict( request.data )
    query_parameters = list( query.keys() )

    for key in valid_keys:
        if key not in query_parameters:
            return Response({
                'status' : "Error",
                'code' : 400,
                'message' : f"Wrong Query Passed {key}"
            }, status=status.HTTP_206_PARTIAL_CONTENT)
    
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email = query['email']

    if re.fullmatch(regex, email):
        username, domain = email.split('@')
        if domain != 'lnmiit.ac.in':
            return Response({
                'status' : "Error",
                'code' : 206,
                'message' : f"Enter a Valid LNMIIT Email Only"
            }, status = status.HTTP_206_PARTIAL_CONTENT)
        name = query['name']

        try:
            recommendObj = Recommend.objects.filter( email = email ).order_by('-created_at')
            if len(list(recommendObj.values())) == 0:
                return Response({
                    'status' : 'Success',
                    'code' : 200,
                    'data' : [],
                    'message' : 'Nothing Found',
                }, status=status.HTTP_200_OK)
            if recommendObj.exists():
                books = []
                for obj in list(recommendObj.values()):
                    bookObj = Book.objects.get( id = obj['book_id'] )
                    books.append({
                        "id" : obj['id'],
                        'qty' : obj['quantity'],
                        "book_id" : bookObj.id,
                        "title" : bookObj.title,
                        "author" : bookObj.author,
                        "subject" : bookObj.subject,
                        "seller" : bookObj.seller.name,
                        "price" : bookObj.expected_price,
                        "medium" : bookObj.medium,
                        "price_denomination" : bookObj.price_denomination,
                    })
                
                return Response({
                        "status" : "Success",
                        "code":200,
                        "type" : "Recommended",
                        "typeCode" : 1,
                        "data":books,
                }, status=status.HTTP_200_OK)
            else:
                    return Response({
                        "status" : "error",
                        "code" : 204,
                        'Message' : "Something Went Wrong, Cant Fetch Recommended"
                    }, status=status.HTTP_204_NO_CONTENT)
        
        except Exception as e:
            print(e)
            return Response({
                'status' : 'Error', 
                'code' : 400,
                'message' : "Something Wrong with Recommend table"
            }, status = status.HTTP_400_BAD_REQUEST)

    else:
        return Response({
            'status' : "Error",
            'code' : 206,
            'message' : f"Enter a Valid Email"
        }, status=status.HTTP_206_PARTIAL_CONTENT)

@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def getOrders(request):
    valid_keys = ['email', 'name']
    query = dict( request.data )
    query_parameters = list( query.keys() )

    for key in valid_keys:
        if key not in query_parameters:
            return Response({
                'status' : "Error",
                'code' : 400,
                'message' : f"Wrong Query Passed {key}"
            }, status=status.HTTP_206_PARTIAL_CONTENT)
    
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email = query['email']

    if re.fullmatch(regex, email):
        username, domain = email.split('@')
        if domain != 'lnmiit.ac.in':
            return Response({
                'status' : "Error",
                'code' : 206,
                'message' : f"Enter a Valid LNMIIT Email Only"
            }, status = status.HTTP_206_PARTIAL_CONTENT)
        name = query['name']

        try:
            ordersObj = Order.objects.filter( email = email ).order_by('-created_at')
            if len(list(ordersObj.values())) == 0:
                return Response({
                    'status' : 'Success',
                    'code' : 200,
                    'data' : [],
                    'message' : 'Nothing Found',
                }, status=status.HTTP_200_OK)
            if ordersObj.exists():
                books = []
                for obj in list(ordersObj.values()):
                    bookObj = Book.objects.get( id = obj['book_id'] )
                    books.append({
                        "id" : obj['id'],
                        'qty' : obj['quantity'],
                        "book_id" : bookObj.id,
                        "title" : bookObj.title,
                        "author" : bookObj.author,
                        "subject" : bookObj.subject,
                        "seller" : bookObj.seller.name,
                        "price" : bookObj.expected_price,
                        "medium" : bookObj.medium,
                        "price_denomination" : bookObj.price_denomination,
                    })
                
                return Response({
                        "status" : "Success",
                        "code":200,
                        "type" : "Purchase",
                        "typeCode" : 2,
                        "data":books,
                }, status=status.HTTP_200_OK)
            
            else:
                    return Response({
                        "status" : "error",
                        "code" : 204,
                        'Message' : "Something Went Wrong, Cant Fetch Purchases"
                    }, status=status.HTTP_204_NO_CONTENT)
        
        except Exception as e:
            print(e)
            return Response({
                'status' : 'Error', 
                'code' : 400,
                'message' : "Something Wrong with Purchases table"
            }, status = status.HTTP_400_BAD_REQUEST)

    else:
        return Response({
            'status' : "Error",
            'code' : 206,
            'message' : f"Enter a Valid Email"
        }, status=status.HTTP_206_PARTIAL_CONTENT)


from django.conf import settings
@api_view(['POST'])
@permission_classes([AllowAny,])
@csrf_exempt
def import_book_data(request):
    if len(request.FILES['files']) == 0:
        return JsonResponse({
            'status' : 'error',
            'code' : 301,
            'message' : " No File Found "
        })
    excel_obj = ExcelFileUpload.objects.create(excel_file_upload = request.FILES['files'])

    df = pd.read_csv(f"{settings.BASE_DIR}/media/{excel_obj.excel_file_upload}", encoding='UTF-8')
    books = df.values.tolist()
    for obj in books:
        title = obj[1]
        author = obj[2]
        subject = obj[3]
        year_of_publication = obj[4]
        edition = obj[5]
        ISBN = obj[6]
        print(ISBN)
        publisher = obj[7]
        seller_id = obj[8]
        if obj[9] is not None:
            price_foreign_currency = obj[9]
        else:
            price_foreign_currency = 0
        
        if obj[10] is not None:
            price_indian_currency = obj[10]
        else:
            price_indian_currency = 0
        link = obj[11]
        book = Book()
        seller = Seller.objects.get(id = seller_id)
        book.title = title
        book.subject = subject
        book.author = author
        book.seller = seller
        book.year_of_publication = year_of_publication
        book.edition = edition
        book.ISBN = ISBN
        book.publisher = publisher
        book.price_indian_currency = price_indian_currency
        book.price_foreign_currency = price_foreign_currency
        book.link = link
        if obj[12] == 'electronic':
            book.medium = obj[12]
        book.save()

        print(f'Book Added {book.id}')

    return JsonResponse({
        'status':'success',
        'code' : 200,
    })


@api_view(['GET'])
@csrf_exempt
def getAdminCount(request, pk):
    if pk is not None:
        if pk > 2 and pk < 1:
            return Response({
                'Status' : "Error",
                'code' : 404,
                'message' : "This type doesnt Exist"
            }, status=status.HTTP_200_OK)
        if pk == 1:
            object = Recommend.objects.all()
        else :
            object = Order.objects.all()
        
        paperback = 0
        ebooks = 0
        cost = 0

        for obj in object:
            if obj.book.medium == "PAPERBACK":
                paperback += 1
                if obj.book.price_denomination == 'INR' or obj.book.price_denomination == 'Rs':
                    cost += obj.book.expected_price
            
            elif obj.book.medium == "electronic":
                ebooks += 1
                if obj.book.price_denomination == 'INR' or obj.book.price_denomination == 'Rs':
                    cost += obj.book.expected_price

        return JsonResponse({
            "status" : "Success",
            "code" : 200,
            "data" : {
                "count" : len( object.values() ),
                "paperbacks" : paperback,
                "ebooks" : ebooks,
                "cost" : cost
            }
        })

@api_view(['GET'])
@csrf_exempt
def bookAction(request, pk):
    if pk > 2 and pk < 1:
            return Response({
                'Status' : "Error",
                'code' : 404,
                'message' : "This type doesnt Exist"
            }, status=status.HTTP_200_OK)
    
    if pk == 1:
        object = Recommend.objects.all()

    else:
        object = Order.objects.all()

    count = [0, 0, 0, 0, 0, 0, 0]

    for obj in object:
        if obj.book.subject == 'CSE':
            count[0] += 1
        elif obj.book.subject == 'CCE':
            count[1] += 1
        elif obj.book.subject == 'ECE':
            count[2] += 1
        elif obj.book.subject == 'MME':
            count[3] += 1
        elif obj.book.subject == 'Physics':
            count[4] += 1
        elif obj.book.subject == 'HSS':
            count[5] += 1
        elif obj.book.subject == 'Mathematics':
            count[6] += 1
    print(count)
    return JsonResponse({
        "status" : "Success",
        "code" : 200,
        "data" : {
            "count" : count
        }
    })


def booksActionsSeller(request, sellerid, type):
    if sellerid > 9 or sellerid < 1:
        return JsonResponse({
            "status" : "error",
            "code" : 404,
            "message" : "Seller ID Exceeded Permissible range"
        })
    if type > 2 or type < 1:
        return JsonResponse({
            "status" : "error",
            "code" : 404,
            "message" : "Type Exceeded Permissible range"
        })
    
    try:
        sellerObj = Seller.objects.get( id = sellerid )
    except:
        return JsonResponse({
            "status" : "error",
            "code" : 404,
            'Message' : "Seller Doesnt Exist"
        })
    
    if type == 1:
        object = Recommend.objects.filter( seller = sellerObj )
    else:
        object = Order.objects.filter( seller = sellerObj )

    
    count = [0, 0, 0, 0, 0, 0, 0]

    for obj in object:
        if obj.book.subject == 'CSE':
            count[0] += 1
        elif obj.book.subject == 'CCE':
            count[1] += 1
        elif obj.book.subject == 'ECE':
            count[2] += 1
        elif obj.book.subject == 'MME':
            count[3] += 1
        elif obj.book.subject == 'Physics':
            count[4] += 1
        elif obj.book.subject == 'HSS':
            count[5] += 1
        elif obj.book.subject == 'Mathematics':
            count[6] += 1
    return JsonResponse({
        "status" : "Success",
        "code" : 200,
        "data" : {
            "count" : count
        }
    })






def recommendExcelApiAll(request, type):
    if type == 1:
        object = Recommend.objects.all()
    else :
        object = Order.objects.all()
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Type"
    ws["D1"] = "Email"
    ws["E1"] = "Date"
    ws["F1"] = "Title"
    ws["G1"] = "Subject"
    ws["H1"] = "Author"
    ws["I1"] = "Year of Publication"
    ws["J1"] = "Seller"
    ws["K1"] = "Publisher"
    ws["L1"] = "ISBN"
    ws["M1"] = "Medium"
    ws["N1"] = "Price_foreign_currency"
    ws["O1"] = "Price_indian_currency"
    ws["P1"] = "Denomination"
    ws["Q1"] = "expected_price"
    ws["R1"] = 'Quantity'
    ws["S1"] = "Book Link"

    row = 2

    for obj in object:
        ws["A{}".format(row)] = obj.id
        ws["B{}".format(row)] = obj.name
        ws["C{}".format(row)] = obj.usertype.name
        ws["D{}".format(row)] = obj.email
        ws["E{}".format(row)] = obj.created_at.date()
        ws["F{}".format(row)] = obj.book.title
        ws["G{}".format(row)] = obj.book.subject
        ws["H{}".format(row)] = obj.book.author
        ws["I{}".format(row)] = obj.book.year_of_publication
        ws["J{}".format(row)] = obj.book.seller.name
        ws["K{}".format(row)] = obj.book.publisher
        ws["L{}".format(row)] = obj.book.ISBN
        ws["M{}".format(row)] = obj.book.medium
        ws["N{}".format(row)] = obj.book.price_foreign_currency
        ws["O{}".format(row)] = obj.book.price_indian_currency
        ws["P{}".format(row)] = obj.book.price_denomination
        ws["Q{}".format(row)] = obj.book.expected_price
        ws["R{}".format(row)] = obj.quantity
        ws["S{}".format(row)] = obj.book.link
        row += 1

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Recommended_List_All_Books.xlsx"

    print("File Created")
    return response

def recommendExcelApiSeller(request, type, seller):
    sellerobj = Seller.objects.get(id = seller)
    if type == 1:
        object = Recommend.objects.filter( seller = sellerobj )
    else :
        object = Order.objects.filter( seller = sellerobj )
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Type"
    ws["D1"] = "Email"
    ws["E1"] = "Date"
    ws["F1"] = "Title"
    ws["G1"] = "Subject"
    ws["H1"] = "Author"
    ws["I1"] = "Year of Publication"
    ws["J1"] = "Seller"
    ws["K1"] = "Publisher"
    ws["L1"] = "ISBN"
    ws["M1"] = "Medium"
    ws["N1"] = "Price_foreign_currency"
    ws["O1"] = "Price_indian_currency"
    ws["P1"] = "Denomination"
    ws["Q1"] = "expected_price"
    ws["R1"] = 'Quantity'
    ws["S1"] = "Book Link"

    row = 2

    for obj in object:
        ws["A{}".format(row)] = obj.id
        ws["B{}".format(row)] = obj.name
        ws["C{}".format(row)] = obj.usertype.name
        ws["D{}".format(row)] = obj.email
        ws["E{}".format(row)] = obj.created_at.date()
        ws["F{}".format(row)] = obj.book.title
        ws["G{}".format(row)] = obj.book.subject
        ws["H{}".format(row)] = obj.book.author
        ws["I{}".format(row)] = obj.book.year_of_publication
        ws["J{}".format(row)] = obj.book.seller.name
        ws["K{}".format(row)] = obj.book.publisher
        ws["L{}".format(row)] = obj.book.ISBN
        ws["M{}".format(row)] = obj.book.medium
        ws["N{}".format(row)] = obj.book.price_foreign_currency
        ws["O{}".format(row)] = obj.book.price_indian_currency
        ws["P{}".format(row)] = obj.book.price_denomination
        ws["Q{}".format(row)] = obj.book.expected_price
        ws["R{}".format(row)] = obj.quantity
        ws["S{}".format(row)] = obj.book.link
        row += 1

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Recommended_List_Seller_Books.xlsx"

    print("File Created")
    return response

def recommendExcelApiSubject(request, type, subject):
    if type == 1:
        object = Recommend.objects.all()
    else :
        object = Order.objects.all()
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Type"
    ws["D1"] = "Email"
    ws["E1"] = "Date"
    ws["F1"] = "Title"
    ws["G1"] = "Subject"
    ws["H1"] = "Author"
    ws["I1"] = "Year of Publication"
    ws["J1"] = "Seller"
    ws["K1"] = "Publisher"
    ws["L1"] = "ISBN"
    ws["M1"] = "Medium"
    ws["N1"] = "Price_foreign_currency"
    ws["O1"] = "Price_indian_currency"
    ws["P1"] = "Denomination"
    ws["Q1"] = "expected_price"
    ws["R1"] = 'Quantity'
    ws["S1"] = "Book Link"

    row = 2

    for obj in object:
        if obj.book.subject.lower() == subject.lower():
            ws["A{}".format(row)] = obj.id
            ws["B{}".format(row)] = obj.name
            ws["C{}".format(row)] = obj.usertype.name
            ws["D{}".format(row)] = obj.email
            ws["E{}".format(row)] = obj.created_at.date()
            ws["F{}".format(row)] = obj.book.title
            ws["G{}".format(row)] = obj.book.subject
            ws["H{}".format(row)] = obj.book.author
            ws["I{}".format(row)] = obj.book.year_of_publication
            ws["J{}".format(row)] = obj.book.seller.name
            ws["K{}".format(row)] = obj.book.publisher
            ws["L{}".format(row)] = obj.book.ISBN
            ws["M{}".format(row)] = obj.book.medium
            ws["N{}".format(row)] = obj.book.price_foreign_currency
            ws["O{}".format(row)] = obj.book.price_indian_currency
            ws["P{}".format(row)] = obj.book.price_denomination
            ws["Q{}".format(row)] = obj.book.expected_price
            ws["R{}".format(row)] = obj.quantity
            ws["S{}".format(row)] = obj.book.link
            row += 1

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Recommended_List_Seller_And_Subject_Books.xlsx"

    print("File Created")
    return response

def recommendExcelApiSellerAndSubject(request, type, seller, subject):
    sellerobj = Seller.objects.get(id = seller)
    if type == 1:
        object = Recommend.objects.filter( seller = sellerobj )
    else :
        object = Order.objects.filter( seller = sellerobj )
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Type"
    ws["D1"] = "Email"
    ws["E1"] = "Date"
    ws["F1"] = "Title"
    ws["G1"] = "Subject"
    ws["H1"] = "Author"
    ws["I1"] = "Year of Publication"
    ws["J1"] = "Seller"
    ws["K1"] = "Publisher"
    ws["L1"] = "ISBN"
    ws["M1"] = "Medium"
    ws["N1"] = "Price_foreign_currency"
    ws["O1"] = "Price_indian_currency"
    ws["P1"] = "Denomination"
    ws["Q1"] = "expected_price"
    ws["R1"] = 'Quantity'
    ws["S1"] = "Book Link"

    row = 2

    for obj in object:
        if obj.book.subject.lower() == subject.lower():
            ws["A{}".format(row)] = obj.id
            ws["B{}".format(row)] = obj.name
            ws["C{}".format(row)] = obj.usertype.name
            ws["D{}".format(row)] = obj.email
            ws["E{}".format(row)] = obj.created_at.date()
            ws["F{}".format(row)] = obj.book.title
            ws["G{}".format(row)] = obj.book.subject
            ws["H{}".format(row)] = obj.book.author
            ws["I{}".format(row)] = obj.book.year_of_publication
            ws["J{}".format(row)] = obj.book.seller.name
            ws["K{}".format(row)] = obj.book.publisher
            ws["L{}".format(row)] = obj.book.ISBN
            ws["M{}".format(row)] = obj.book.medium
            ws["N{}".format(row)] = obj.book.price_foreign_currency
            ws["O{}".format(row)] = obj.book.price_indian_currency
            ws["P{}".format(row)] = obj.book.price_denomination
            ws["Q{}".format(row)] = obj.book.expected_price
            ws["R{}".format(row)] = obj.quantity
            ws["S{}".format(row)] = obj.book.link
            row += 1

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Recommended_List_Seller_And_Subject_Books.xlsx"

    print("File Created")
    return response

